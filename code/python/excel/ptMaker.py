from openpyxl import load_workbook
import os
templatePath = r'PT仕様書【】_Template.xlsx'
ssPath = r'D:\ProjectDocument\寿司\使用量入力\SS_画面設計書【S03T21F03P01 日別使用量入力（入力）】.xlsx'

templateWb = load_workbook(filename=templatePath)
ssWb = load_workbook(filename=ssPath)

class PTMaker(object):
    def __init__(self, templateWb, ssWb):
        self.templateWb = templateWb
        self.ssWb = ssWb

    def item_define(self):
        uiSheet = self.templateWb['画面項目定義']
        ssSheet = self.ssWb['画面項目定義']
        ssSheet["U10"].value = uiSheet["M15"].value

    def close(self):
        print(ssPath)
        self.ssWb.save(filename=ssPath)
        self.templateWb.close()
        self.ssWb.close()

ssMaker = PTMaker(templateWb, ssWb)
ssMaker.item_define()
ssWb.save("./test.xlsx")
ssMaker.close()


