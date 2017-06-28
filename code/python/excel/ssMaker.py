from openpyxl import load_workbook
import os
uiPath = r'D:\ProjectDocument\寿司\使用量入力\02_050_日別使用量入力_02_画面設計書_01_日別使用量入力（入力）.xlsx'
ssPath = r'D:\ProjectDocument\寿司\使用量入力\SS_画面設計書【S03T21F03P01 日別使用量入力（入力）】.xlsx'

uiWb = load_workbook(filename=uiPath)
ssWb = load_workbook(filename=ssPath)

class SSMaker(object):
    def __init__(self, uiWb, ssWb):
        self.uiWb = uiWb
        self.ssWb = ssWb

    def item_define(self):
        uiSheet = self.uiWb['画面項目定義']
        ssSheet = self.ssWb['画面項目定義']
        ssSheet["U10"].value = uiSheet["M15"].value

    def close(self):
        print(ssPath)
        self.ssWb.save(filename=ssPath)
        self.uiWb.close()
        self.ssWb.close()

ssMaker = SSMaker(uiWb, ssWb)
ssMaker.item_define()
ssWb.save("./test.xlsx")
ssMaker.close()
